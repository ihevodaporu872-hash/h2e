"""Excel file parser (.xlsx, .xls)."""

from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .base import (
    BaseParser,
    DocumentContent,
    ParsedTable,
    ParserErrorType,
    ParserResult,
)


class ExcelParser(BaseParser):
    """Parser for Excel files (.xlsx, .xls)."""

    SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]

    def parse(self, file_path: Path) -> ParserResult:
        """Parse Excel file and extract content."""
        if not file_path.exists():
            return self._create_error_result(
                file_path,
                ParserErrorType.FILE_NOT_FOUND,
                f"File not found: {file_path}",
            )

        result = ParserResult(
            file_path=file_path,
            file_type=file_path.suffix.lower(),
            success=True,
        )

        try:
            if file_path.suffix.lower() == ".xlsx":
                self._parse_xlsx(file_path, result)
            else:
                self._parse_xls(file_path, result)

        except InvalidFileException:
            return self._create_error_result(
                file_path,
                ParserErrorType.FILE_CORRUPTED,
                "Invalid or corrupted Excel file.",
            )
        except Exception as e:
            error_msg = str(e).lower()
            if "password" in error_msg or "encrypted" in error_msg:
                return self._create_error_result(
                    file_path,
                    ParserErrorType.PASSWORD_PROTECTED,
                    "Excel file is password protected.",
                )
            return self._create_error_result(
                file_path,
                ParserErrorType.FILE_CORRUPTED,
                f"Failed to parse Excel file: {e}",
            )

        return result

    def _parse_xlsx(self, file_path: Path, result: ParserResult) -> None:
        """Parse .xlsx file using openpyxl for better metadata."""
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            result.metadata["sheet_names"] = workbook.sheetnames
            result.total_pages = len(workbook.sheetnames)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                self._process_sheet(sheet_name, sheet, result)

            workbook.close()

        except Exception:
            # Fallback to pandas
            self._parse_with_pandas(file_path, result)

    def _parse_xls(self, file_path: Path, result: ParserResult) -> None:
        """Parse .xls file using pandas."""
        self._parse_with_pandas(file_path, result)

    def _parse_with_pandas(self, file_path: Path, result: ParserResult) -> None:
        """Fallback parsing using pandas."""
        excel_file = pd.ExcelFile(file_path)
        result.metadata["sheet_names"] = excel_file.sheet_names
        result.total_pages = len(excel_file.sheet_names)

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            self._process_dataframe(sheet_name, df, result)

    def _process_sheet(self, sheet_name: str, sheet: Any, result: ParserResult) -> None:
        """Process a single Excel sheet."""
        rows_data = []
        text_content = []

        for row in sheet.iter_rows():
            row_values = []
            for cell in row:
                value = cell.value
                if value is not None:
                    str_value = str(value).strip()
                    row_values.append(str_value)
                else:
                    row_values.append("")

            # Skip completely empty rows
            if any(row_values):
                rows_data.append(row_values)

        if not rows_data:
            return

        # Detect if this looks like a table (has headers) or free text
        is_table = self._detect_table_structure(rows_data)

        if is_table and len(rows_data) >= 2:
            parsed_table = self._create_table(rows_data, sheet_name)
            if parsed_table:
                result.tables.append(parsed_table)
        else:
            # Treat as text content
            text = "\n".join(" | ".join(filter(None, row)) for row in rows_data)
            result.content_sections.append(
                DocumentContent(
                    text=text,
                    section_title=sheet_name,
                    metadata={"sheet": sheet_name},
                )
            )

    def _process_dataframe(
        self,
        sheet_name: str,
        df: pd.DataFrame,
        result: ParserResult,
    ) -> None:
        """Process pandas DataFrame from sheet."""
        # Drop completely empty rows and columns
        df = df.dropna(how="all").dropna(axis=1, how="all")

        if df.empty:
            return

        # Convert to list of lists
        rows_data = df.fillna("").astype(str).values.tolist()

        if len(rows_data) < 2:
            return

        is_table = self._detect_table_structure(rows_data)

        if is_table:
            parsed_table = self._create_table(rows_data, sheet_name)
            if parsed_table:
                result.tables.append(parsed_table)
        else:
            text = "\n".join(" | ".join(filter(None, row)) for row in rows_data)
            result.content_sections.append(
                DocumentContent(
                    text=text,
                    section_title=sheet_name,
                    metadata={"sheet": sheet_name},
                )
            )

    def _detect_table_structure(self, rows: list[list[str]]) -> bool:
        """Detect if rows represent a structured table."""
        if len(rows) < 2:
            return False

        # Check if first row looks like headers
        first_row = rows[0]
        data_rows = rows[1:]

        # Headers typically:
        # 1. Are shorter than data
        # 2. Don't contain numbers primarily
        # 3. Have consistent column count

        header_lengths = [len(cell) for cell in first_row if cell]
        if not header_lengths:
            return False

        avg_header_length = sum(header_lengths) / len(header_lengths)

        # Check data consistency
        non_empty_counts = [sum(1 for cell in row if cell) for row in data_rows]
        if not non_empty_counts:
            return False

        # If most rows have similar column usage, it's likely a table
        avg_cols = sum(non_empty_counts) / len(non_empty_counts)

        return avg_cols >= 2 and avg_header_length < 100

    def _create_table(
        self,
        rows: list[list[str]],
        sheet_name: str,
    ) -> Optional[ParsedTable]:
        """Create ParsedTable from rows."""
        if len(rows) < 2:
            return None

        # Normalize column count
        max_cols = max(len(row) for row in rows)
        normalized = [row + [""] * (max_cols - len(row)) for row in rows]

        return ParsedTable(
            headers=normalized[0],
            rows=normalized[1:],
            confidence=1.0,
        )
