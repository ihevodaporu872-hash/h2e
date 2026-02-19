"""Excel BOQ generator with template support."""

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..assembly.assembler import AssembledBOQ, BOQSection
from ..extraction.extractor import ExtractedItem


@dataclass
class ColumnDef:
    """Column definition for BOQ output."""

    name: str
    width: int = 15
    type: str = "text"  # text, number, currency, formula
    decimal_places: int = 2
    currency_symbol: str = "$"
    formula: Optional[str] = None
    wrap_text: bool = False


@dataclass
class BOQTemplate:
    """BOQ output template configuration."""

    name: str = "Default Template"
    version: str = "1.0"
    sheet_name: str = "Bill of Quantities"
    start_row: int = 5
    columns: list[ColumnDef] = field(default_factory=list)
    contingency_percent: float = 5.0
    currency_symbol: str = "$"

    @classmethod
    def from_yaml(cls, path: Path) -> "BOQTemplate":
        """Load template from YAML file."""
        with open(path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)

        columns = []
        for col_data in data.get("columns", []):
            columns.append(
                ColumnDef(
                    name=col_data.get("name", ""),
                    width=col_data.get("width", 15),
                    type=col_data.get("type", "text"),
                    decimal_places=col_data.get("decimal_places", 2),
                    currency_symbol=col_data.get("currency_symbol", "$"),
                    formula=col_data.get("formula"),
                    wrap_text=col_data.get("wrap_text", False),
                )
            )

        return cls(
            name=data.get("name", "Custom Template"),
            version=data.get("version", "1.0"),
            sheet_name=data.get("output", {}).get("sheet_name", "Bill of Quantities"),
            start_row=data.get("output", {}).get("start_row", 5),
            columns=columns,
            contingency_percent=data.get("totals", {}).get("contingency", {}).get(
                "percent", 5.0
            ),
            currency_symbol=data.get("currency_symbol", "$"),
        )

    @classmethod
    def default(cls) -> "BOQTemplate":
        """Create default template."""
        return cls(
            columns=[
                ColumnDef(name="Item No.", width=10, type="text"),
                ColumnDef(name="Description", width=60, type="text", wrap_text=True),
                ColumnDef(name="Unit", width=10, type="text"),
                ColumnDef(name="Quantity", width=12, type="number", decimal_places=2),
                ColumnDef(name="Rate", width=15, type="currency", decimal_places=2),
                ColumnDef(
                    name="Amount",
                    width=15,
                    type="formula",
                    formula="=D{row}*E{row}",
                    decimal_places=2,
                ),
            ]
        )


class ExcelGenerator:
    """Generates formatted Excel BOQ documents.

    Example:
        >>> generator = ExcelGenerator()
        >>> generator.generate(assembled_boq, Path("output.xlsx"))
    """

    # Style definitions
    STYLES = {
        "title": {
            "font": Font(size=16, bold=True),
            "alignment": Alignment(horizontal="center"),
        },
        "subtitle": {
            "font": Font(size=12, bold=True),
        },
        "header": {
            "font": Font(size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        },
        "section_header": {
            "font": Font(size=11, bold=True),
            "fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        },
        "data": {
            "font": Font(size=10),
            "alignment": Alignment(vertical="top"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        },
        "subtotal": {
            "font": Font(size=10, bold=True),
            "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        },
        "total": {
            "font": Font(size=11, bold=True),
            "fill": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
            "border": Border(
                left=Side(style="medium"),
                right=Side(style="medium"),
                top=Side(style="medium"),
                bottom=Side(style="medium"),
            ),
        },
    }

    def __init__(self, template: Optional[BOQTemplate] = None):
        """Initialize generator.

        Args:
            template: BOQ template to use. Uses default if not provided.
        """
        self.template = template or BOQTemplate.default()

    def generate(
        self,
        boq: AssembledBOQ,
        output_path: Path,
        include_specs: bool = True,
    ) -> Path:
        """Generate Excel BOQ document.

        Args:
            boq: Assembled BOQ to output.
            output_path: Path for output file.
            include_specs: Whether to include specifications column.

        Returns:
            Path to generated file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = self.template.sheet_name

        current_row = 1

        # Write header
        current_row = self._write_header(ws, boq, current_row)

        # Write column headers
        current_row = self._write_column_headers(ws, current_row)

        data_start_row = current_row

        # Write sections
        for section in boq.sections:
            current_row = self._write_section(ws, section, current_row)

        # Write unclassified items
        if boq.unclassified_items:
            current_row = self._write_unclassified(ws, boq.unclassified_items, current_row)

        data_end_row = current_row - 1

        # Write totals
        current_row = self._write_totals(ws, boq, current_row, data_start_row, data_end_row)

        # Set column widths
        self._set_column_widths(ws)

        # Freeze panes
        ws.freeze_panes = f"A{self.template.start_row + 1}"

        # Save
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return output_path

    def _write_header(self, ws: Any, boq: AssembledBOQ, row: int) -> int:
        """Write document header."""
        # Title
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = "BILL OF QUANTITIES"
        cell.font = self.STYLES["title"]["font"]
        cell.alignment = self.STYLES["title"]["alignment"]
        row += 1

        # Project name
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = f"Project: {boq.project_name}"
        cell.font = self.STYLES["subtitle"]["font"]
        row += 1

        # Date
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
        cell.font = Font(size=10)
        row += 2  # Extra blank row

        return row

    def _write_column_headers(self, ws: Any, row: int) -> int:
        """Write column headers."""
        for col_idx, col_def in enumerate(self.template.columns, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = col_def.name

            # Apply header style
            cell.font = self.STYLES["header"]["font"]
            cell.fill = self.STYLES["header"]["fill"]
            cell.alignment = self.STYLES["header"]["alignment"]
            cell.border = self.STYLES["header"]["border"]

        return row + 1

    def _write_section(self, ws: Any, section: BOQSection, row: int) -> int:
        """Write a BOQ section."""
        num_cols = len(self.template.columns)

        # Section header
        ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
        cell = ws[f"A{row}"]
        cell.value = section.name

        # Apply section header style
        cell.font = self.STYLES["section_header"]["font"]
        cell.fill = self.STYLES["section_header"]["fill"]
        cell.border = self.STYLES["section_header"]["border"]

        row += 1

        # Section items
        for item in section.items:
            row = self._write_item(ws, item, row)

        # Section subtotal
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws[f"A{row}"]
        cell.value = f"{section.name} - Subtotal"
        cell.font = self.STYLES["subtotal"]["font"]
        cell.fill = self.STYLES["subtotal"]["fill"]
        cell.border = self.STYLES["subtotal"]["border"]

        # Subtotal amount
        amount_col = self._get_amount_column()
        subtotal_cell = ws.cell(row=row, column=amount_col)
        subtotal_cell.value = section.subtotal
        subtotal_cell.number_format = f'{self.template.currency_symbol}#,##0.00'
        subtotal_cell.font = self.STYLES["subtotal"]["font"]
        subtotal_cell.fill = self.STYLES["subtotal"]["fill"]
        subtotal_cell.border = self.STYLES["subtotal"]["border"]

        return row + 2  # Extra blank row after section

    def _write_item(self, ws: Any, item: ExtractedItem, row: int) -> int:
        """Write a single BOQ item."""
        col_mapping = {
            "Item No.": item.item_number,
            "Description": item.description,
            "Unit": item.unit or "",
            "Quantity": item.quantity,
            "Rate": item.rate,
            "Amount": item.amount if item.amount else (
                item.quantity * item.rate if item.quantity and item.rate else None
            ),
        }

        for col_idx, col_def in enumerate(self.template.columns, 1):
            cell = ws.cell(row=row, column=col_idx)
            value = col_mapping.get(col_def.name, "")

            # Handle formula columns
            if col_def.type == "formula" and col_def.formula:
                cell.value = col_def.formula.replace("{row}", str(row))
            else:
                cell.value = value

            # Apply formatting
            cell.font = self.STYLES["data"]["font"]
            cell.border = self.STYLES["data"]["border"]

            if col_def.wrap_text:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = self.STYLES["data"]["alignment"]

            # Number formatting
            if col_def.type == "currency" and value is not None:
                cell.number_format = f'{self.template.currency_symbol}#,##0.{"0" * col_def.decimal_places}'
            elif col_def.type == "number" and value is not None:
                cell.number_format = f'#,##0.{"0" * col_def.decimal_places}'

        return row + 1

    def _write_unclassified(
        self,
        ws: Any,
        items: list[ExtractedItem],
        row: int,
    ) -> int:
        """Write unclassified items section."""
        num_cols = len(self.template.columns)

        # Section header
        ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
        cell = ws[f"A{row}"]
        cell.value = "OTHER ITEMS"
        cell.font = self.STYLES["section_header"]["font"]
        cell.fill = self.STYLES["section_header"]["fill"]

        row += 1

        for i, item in enumerate(items, 1):
            item.item_number = f"X.{i}"
            row = self._write_item(ws, item, row)

        return row + 1

    def _write_totals(
        self,
        ws: Any,
        boq: AssembledBOQ,
        row: int,
        data_start: int,
        data_end: int,
    ) -> int:
        """Write totals section."""
        amount_col = self._get_amount_column()
        amount_letter = get_column_letter(amount_col)

        # Subtotal
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = "SUBTOTAL"
        cell.font = self.STYLES["total"]["font"]
        cell.fill = self.STYLES["total"]["fill"]
        cell.border = self.STYLES["total"]["border"]

        subtotal_cell = ws.cell(row=row, column=amount_col)
        subtotal_cell.value = boq.subtotal
        subtotal_cell.number_format = f'{self.template.currency_symbol}#,##0.00'
        subtotal_cell.font = self.STYLES["total"]["font"]
        subtotal_cell.fill = self.STYLES["total"]["fill"]
        subtotal_cell.border = self.STYLES["total"]["border"]

        subtotal_row = row
        row += 1

        # Contingency
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = f"Contingency ({self.template.contingency_percent}%)"
        cell.font = self.STYLES["subtotal"]["font"]
        cell.fill = self.STYLES["subtotal"]["fill"]
        cell.border = self.STYLES["subtotal"]["border"]

        cont_cell = ws.cell(row=row, column=amount_col)
        cont_cell.value = boq.contingency
        cont_cell.number_format = f'{self.template.currency_symbol}#,##0.00'
        cont_cell.font = self.STYLES["subtotal"]["font"]
        cont_cell.fill = self.STYLES["subtotal"]["fill"]
        cont_cell.border = self.STYLES["subtotal"]["border"]

        contingency_row = row
        row += 1

        # Grand Total
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = "GRAND TOTAL"
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.border = self.STYLES["total"]["border"]

        total_cell = ws.cell(row=row, column=amount_col)
        total_cell.value = boq.grand_total
        total_cell.number_format = f'{self.template.currency_symbol}#,##0.00'
        total_cell.font = Font(size=12, bold=True, color="FFFFFF")
        total_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        total_cell.border = self.STYLES["total"]["border"]

        return row + 1

    def _get_amount_column(self) -> int:
        """Get the column index for amount."""
        for i, col in enumerate(self.template.columns, 1):
            if col.name.lower() == "amount":
                return i
        return len(self.template.columns)

    def _set_column_widths(self, ws: Any) -> None:
        """Set column widths from template."""
        for i, col_def in enumerate(self.template.columns, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = col_def.width
